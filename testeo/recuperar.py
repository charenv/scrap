"""Reclama las filas que se perdieron en las paginas que vinieron incompletas.

El buscador sirve 10 resoluciones por pagina. Al cosechar, 415 paginas
devolvieron menos y nadie se entero, porque la cosecha termina "bien" igual:
solo con menos filas. Este script vuelve a pedir exactamente esas paginas -no
las 20 923- y añade al corpus las resoluciones cuyo uuid no estuviera ya.

No borra ni reescribe nada existente: produce un Excel nuevo al lado. Las filas
recuperadas se colocan al final de su propia pagina, para que el fichero siga
agrupado por especialidad y pagina como el original.

Uso:
    python3 recuperar.py                       # todas las especialidades
    python3 recuperar.py --especialidad Civil
    python3 recuperar.py --listar              # solo dice que pediria
"""

import argparse
import json
import sys
import time
from collections import defaultdict
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import juris_parse as jp
from juris_client import ESPECIALIDADES, ClienteJuris

SALIDA = Path("/home/charen/corpus-corte-suprema-2026-08-15/scrap/salida")
CORPUS = SALIDA / "todo-corte-suprema.xlsx"
TAMANO_PAGINA = 10

COLUMNAS = [
    "Casacion/Apelacion",
    "Nro Expediente",
    "Pretension/Delito",
    "Tipo de Resolucion",
    "Fecha de Resolucion",
    "Sala Suprema",
    "Norma de Derecho Interno (Articulo)",
    "Sumilla",
    "Palabras Clave",
    "Link Resolucion",
    "Especialidad",
    "Pagina",
]

# slug de carpeta -> nombre de especialidad tal y como lo espera el cliente
SLUGS = {
    __import__("re").sub(r"[^a-z0-9]+", "-",
                         __import__("unicodedata").normalize("NFKD", e)
                         .encode("ascii", "ignore").decode().lower()).strip("-"): e
    for e in ESPECIALIDADES
}


def paginas_incompletas():
    """{especialidad: [paginas]} a partir de los tramos ya guardados.

    Se excluye la ultima pagina de cada listado, que legitimamente trae menos.
    """
    por_esp = defaultdict(lambda: defaultdict(int))
    for fichero in sorted(SALIDA.glob("*/tramos/*.json")):
        especialidad = SLUGS.get(fichero.parent.parent.name)
        if not especialidad:
            continue
        for fila in json.loads(fichero.read_text(encoding="utf-8"))["filas"]:
            por_esp[especialidad][fila["Pagina"]] += 1

    pendientes = {}
    for especialidad, conteo in por_esp.items():
        ultima = max(conteo)
        cortas = sorted(p for p, n in conteo.items() if n < TAMANO_PAGINA and p != ultima)
        if cortas:
            pendientes[especialidad] = cortas
    return pendientes


CADA = 20  # cada cuantas paginas se informa del avance
REGISTRO = SALIDA / "recuperacion"  # avance en disco, para poder reanudar


def _diario(especialidad):
    """(fichero de filas, fichero de paginas hechas) de una especialidad."""
    slug = next(s for s, e in SLUGS.items() if e == especialidad)
    REGISTRO.mkdir(parents=True, exist_ok=True)
    return REGISTRO / f"{slug}.jsonl", REGISTRO / f"{slug}-hechas.txt"


def ya_hecho(especialidad):
    """Lo recuperado en corridas anteriores: (filas, paginas ya pedidas)."""
    filas_f, hechas_f = _diario(especialidad)
    filas = []
    if filas_f.exists():
        filas = [json.loads(l) for l in filas_f.read_text(encoding="utf-8").splitlines() if l]
    hechas = set()
    if hechas_f.exists():
        hechas = {int(l) for l in hechas_f.read_text(encoding="utf-8").split() if l.strip()}
    return filas, hechas


def reclamar(especialidad, paginas, conocidos, pausa):
    """Re-pide las paginas y devuelve las filas cuyo uuid no estuviera ya.

    Cada pagina se anota en disco nada mas procesarla, asi que una corrida
    cortada se retoma donde iba en vez de repetir media hora de peticiones.
    Informa cada CADA paginas: Civil son 388 peticiones seguidas y sin esto la
    corrida parece colgada durante todo ese rato.
    """
    filas_f, hechas_f = _diario(especialidad)
    previas, hechas = ya_hecho(especialidad)
    for fila in previas:
        conocidos.add(fila["uuid"])
    pendientes = [p for p in paginas if p not in hechas]
    if hechas:
        print(f"    reanudando: {len(hechas)} paginas ya hechas, "
              f"{len(previas)} filas ya recuperadas, quedan {len(pendientes)}", flush=True)
    if not pendientes:
        return previas, 0, 0

    cliente = ClienteJuris(especialidad=especialidad, pausa=pausa)
    cliente.pagina_con_reintentos(1, primera=True)

    inicio = time.monotonic()
    nuevas, revisadas, fallos = list(previas), 0, 0
    with filas_f.open("a", encoding="utf-8") as diario, hechas_f.open("a", encoding="utf-8") as marcas:
        for i, numero in enumerate(pendientes, start=1):
            try:
                html = cliente.pagina_con_reintentos(numero)
            except Exception as e:
                fallos += 1
                print(f"    p{numero}: FALLO {type(e).__name__}: {e}", flush=True)
                continue
            revisadas += 1
            for fila in jp.parsear_resultados(html, especialidad, numero):
                if fila["uuid"] not in conocidos:
                    conocidos.add(fila["uuid"])
                    nuevas.append(fila)
                    diario.write(json.dumps(fila, ensure_ascii=False) + "\n")
            marcas.write(f"{numero}\n")
            diario.flush()
            marcas.flush()

            if i % CADA == 0 or i == len(pendientes):
                transcurrido = time.monotonic() - inicio
                queda = transcurrido / i * (len(pendientes) - i)
                print(f"    {i}/{len(pendientes)} paginas  {len(nuevas)} filas nuevas  "
                      f"{transcurrido/i:.1f}s/pag  quedan ~{queda/60:.0f} min", flush=True)
    return nuevas, revisadas, fallos


def escribir(df, destino):
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    libro = Workbook(write_only=True)
    hoja = libro.create_sheet("Resoluciones")
    hoja.freeze_panes = "A2"
    anchos = [18, 16, 34, 20, 16, 34, 30, 70, 34, 62, 22, 8]
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho

    negrita = Font(bold=True)
    cabecera = []
    for col in df.columns:
        celda = WriteOnlyCell(hoja, value=col)
        celda.font = negrita
        celda.alignment = Alignment(vertical="center")
        cabecera.append(celda)
    hoja.append(cabecera)
    for fila in df.itertuples(index=False, name=None):
        hoja.append([None if pd.isna(v) else v for v in fila])
    libro.save(destino)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--especialidad", choices=sorted(ESPECIALIDADES))
    ap.add_argument("--pausa", type=float, default=2.0)
    ap.add_argument("--listar", action="store_true")
    args = ap.parse_args()

    pendientes = paginas_incompletas()
    if args.especialidad:
        pendientes = {k: v for k, v in pendientes.items() if k == args.especialidad}

    total = sum(len(v) for v in pendientes.values())
    print(f"{'ESPECIALIDAD':36} {'PAGINAS INCOMPLETAS':>20}")
    for especialidad, paginas in sorted(pendientes.items(), key=lambda x: -len(x[1])):
        print(f"{especialidad:36} {len(paginas):>20,}")
    print(f"{'TOTAL':36} {total:>20,}   (~{total * args.pausa / 60:.0f} min)")
    if args.listar:
        return 0

    print(f"\nleyendo {CORPUS}")
    corpus = pd.read_excel(CORPUS, engine="openpyxl")
    corpus["_orden"] = range(len(corpus))
    conocidos = set(corpus["Link Resolucion"].str.extract(jp._HREF_DESCARGA, expand=False))
    print(f"  {len(corpus):,} filas, {len(conocidos):,} uuid conocidos\n")

    recuperadas, resumen = [], []
    for especialidad, paginas in sorted(pendientes.items()):
        print(f"{especialidad} ({len(paginas)} paginas)", flush=True)
        nuevas, revisadas, fallos = reclamar(especialidad, paginas, conocidos, args.pausa)
        recuperadas.extend(nuevas)
        resumen.append((especialidad, len(paginas), revisadas, len(nuevas), fallos))
        print(f"  -> {len(nuevas)} filas nuevas ({revisadas} paginas revisadas, {fallos} fallos)\n",
              flush=True)

    print(f"{'ESPECIALIDAD':36} {'PAGS':>6} {'REVIS':>6} {'NUEVAS':>7} {'FALLOS':>7}")
    for fila in resumen:
        print(f"{fila[0]:36} {fila[1]:6d} {fila[2]:6d} {fila[3]:7d} {fila[4]:7d}")
    print(f"\nTOTAL RECUPERADO: {len(recuperadas):,} filas")

    if not recuperadas:
        print("nada que añadir; el corpus se queda igual")
        return 0

    crudo = SALIDA / "recuperadas.json"
    crudo.write_text(json.dumps(recuperadas, ensure_ascii=False), encoding="utf-8")
    print(f"filas crudas -> {crudo}")

    nuevas = pd.DataFrame(recuperadas)[COLUMNAS]
    # Se cuelan al final de su pagina: mismo orden de especialidad y pagina que
    # el corpus, y detras de las filas que ya estaban ahi.
    nuevas["_orden"] = len(corpus)
    orden_esp = {e: i for i, e in enumerate(corpus["Especialidad"].drop_duplicates())}
    junto = pd.concat([corpus, nuevas], ignore_index=True)
    junto["_esp"] = junto["Especialidad"].map(orden_esp).fillna(len(orden_esp))
    junto = junto.sort_values(["_esp", "Pagina", "_orden"], kind="stable")
    junto = junto.drop(columns=["_orden", "_esp"]).reset_index(drop=True)

    destino = SALIDA / "todo-corte-suprema-completo.xlsx"
    print(f"\nescribiendo {destino}")
    escribir(junto[COLUMNAS], destino)
    print(f"  {len(corpus):,} + {len(nuevas):,} = {len(junto):,} filas")
    print(f"  uuid distintos: {junto['Link Resolucion'].str.extract(jp._HREF_DESCARGA, expand=False).nunique():,}")
    print(f"\nahora: python3 limpieza.py {destino}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
