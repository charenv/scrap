"""Normaliza el corpus de la Corte Suprema sin tocar el fichero original.

Que hace, en orden:

  1. Normaliza las categorias que solo difieren en puntuacion o mayusculas
     ('Consulta.' -> 'Consulta', las tres grafias de 'Ejecutoria Suprema').
  2. Quita el prefijo redundante 'Sumilla:' / 'Sumilla.' que el redactor copio
     dentro del campo, y colapsa espacios dobles y saltos de linea.
  3. Marca -no borra- los casi-duplicados: filas identicas en las diez columnas
     de contenido pero con uuid distinto, que son repeticiones del portal.
  4. Añade columnas derivadas utiles (uuid, fecha como fecha, año) y banderas
     de calidad, todas al final para no alterar el orden original.
  5. Calcula 'Archivo PDF': la ruta con la que se guardara cada descarga. Lleva
     el uuid dentro a proposito, porque 'recurso-expediente' colisiona en 14 525
     filas y guardarlas asi sobrescribiria un PDF de cada siete.

Ademas escribe un manifiesto CSV (uuid, url, ruta) listo para alimentar al
descargador, con una fila por resolucion y rutas garantizadas unicas.

Nada se elimina y ninguna fila cambia de sitio: la verificacion final lo
comprueba comparando la secuencia de uuid antes y despues.

Uso:
    python3 limpieza.py                          # rutas por defecto
    python3 limpieza.py ENTRADA.xlsx SALIDA.xlsx
    python3 limpieza.py --solo-informe           # no escribe nada
"""

import argparse
import re
import sys
import unicodedata
from pathlib import Path

import pandas as pd

ENTRADA = Path("/home/charen/corpus-corte-suprema-2026-08-15/scrap/salida/todo-corte-suprema.xlsx")

# Las doce columnas originales, en su orden. Lo que añadimos va detras.
ORIGINALES = [
    "Casacion/Apelacion",
    "Nro Expediente",
    "Pretension/Delito",
    "Tipo de Resolucion",
    "Fecha de Resolucion",
    "Sala Suprema",
    "Norma de Derecho Interno (Articulo)",
    "Sumilla",
    "Palabras Clave",
    "Link Resolucion",
    "Especialidad",
    "Pagina",
]

# Columnas de texto libre: se les limpia el espaciado, no el contenido.
TEXTO = [
    "Pretension/Delito",
    "Norma de Derecho Interno (Articulo)",
    "Sumilla",
    "Palabras Clave",
]

# Columnas que definen "la misma resolucion" para detectar repeticiones del
# portal: todo menos el uuid (que es justo lo que difiere) y la pagina.
CONTENIDO = [c for c in ORIGINALES if c not in ("Link Resolucion", "Pagina")]

_UUID = re.compile(r"uuid=([0-9a-fA-F-]{36})")
# El separador puede ser de varios caracteres ('Sumilla.-', 'Sumilla:-') y la
# etiqueta llega a venir repetida ('Sumilla: Sumilla. El derecho...'), asi que
# se consume el grupo entero tantas veces como aparezca.
_PREFIJO_SUMILLA = re.compile(r"^(?:\s*sumilla\s*[.:–—-]+\s*)+", re.IGNORECASE)
# Lo que se borra de verdad: la etiqueta y/o la puntuacion suelta que el
# redactor dejo delante (': En los casos...', '.- Sumilla: El texto...'). Va en
# un solo patron repetible porque quitar una capa destapa la siguiente.
_LIMPIEZA_INICIAL = re.compile(r"^(?:\s*(?:sumilla\s*)?[.:–—-]+\s*)+", re.IGNORECASE)
_ESPACIOS = re.compile(r"\s+")
_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Sumillas que no son sumillas sino el fallo colado en el campo.
LARGO_MINIMO_SUMILLA = 20

# Cuantos caracteres del uuid se cuelgan del nombre del PDF. Con 8 ya no hay
# ninguna colision en las 208 641 filas; el uuid completo va en el manifiesto.
CORTE_UUID = 8


def _slug(texto):
    """'Contencioso Adm. Laboral' -> 'contencioso-adm-laboral'."""
    if not isinstance(texto, str):
        return "sin-especialidad"
    plano = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", plano.lower()).strip("-") or "sin-especialidad"


def ruta_pdf(recurso, expediente, especialidad, uuid, anio):
    """Ruta relativa del PDF: <especialidad>/<año>/<recurso>-<expediente>-<uuid8>.

    Dos decisiones, las dos por motivos practicos:

    El uuid al final porque 'recurso-expediente' se repite en 12 861 nombres -un
    mismo expediente acumula varias resoluciones- y guardarlos asi perderia
    15 064 PDFs por sobrescritura silenciosa.

    El año en medio porque sin el, penal/ acumula 70 912 ficheros en una sola
    carpeta, que Drive no abre; y porque el corte entre lo que necesita OCR y lo
    que no es por año, asi que cada carpeta-año es un lote de trabajo completo.
    """
    base = _slug(recurso if isinstance(recurso, str) and recurso.strip() else "resolucion")
    nro = re.sub(r"[^0-9-]", "", expediente or "")
    corto = (uuid or "")[:CORTE_UUID]
    nombre = "-".join(p for p in (base, nro, corto) if p)
    carpeta_anio = str(anio) if pd.notna(anio) else "sin-anio"
    return f"{_slug(especialidad)}/{carpeta_anio}/{nombre}.pdf"


def _texto(serie):
    """Colapsa espacios y saltos de linea. Los nulos siguen siendo nulos."""
    limpia = serie.astype("string")
    limpia = limpia.str.replace(_CONTROL, "", regex=True)
    limpia = limpia.str.replace(_ESPACIOS, " ", regex=True).str.strip()
    return limpia.replace("", pd.NA)


def _categoria(serie):
    """Unifica variantes que solo difieren en espacios, punto final o caja.

    Se conserva la grafia mayoritaria como forma canonica en vez de imponer
    una minuscula artificial: 'Consulta.' se convierte en 'Consulta' porque es
    la que aparece 3457 veces, no al reves.
    """
    limpia = _texto(serie)
    clave = (
        limpia.str.normalize("NFKD")
        .str.encode("ascii", "ignore")
        .str.decode("utf-8")
        .str.rstrip(".")
        .str.lower()
    )
    # Forma canonica = la variante mas frecuente de cada clave.
    tabla = (
        pd.DataFrame({"clave": clave, "valor": limpia})
        .dropna()
        .groupby(["clave", "valor"])
        .size()
        .rename("n")
        .reset_index()
        .sort_values(["clave", "n"], ascending=[True, False])
        .drop_duplicates("clave")
        .set_index("clave")["valor"]
    )
    return clave.map(tabla).astype("string")


def limpiar(df):
    """Devuelve una copia normalizada. No modifica df."""
    out = df.copy()

    for col in TEXTO:
        out[col] = _texto(out[col])
    for col in ("Nro Expediente", "Link Resolucion", "Especialidad"):
        out[col] = _texto(out[col])

    # 1. Categorias
    for col in ("Casacion/Apelacion", "Tipo de Resolucion", "Sala Suprema", "Especialidad"):
        out[col] = _categoria(out[col])

    # 2. Prefijo 'Sumilla:' redundante y puntuacion huerfana al inicio
    tenia_prefijo = out["Sumilla"].str.contains(_PREFIJO_SUMILLA, na=False)
    out["Sumilla"] = out["Sumilla"].str.replace(_LIMPIEZA_INICIAL, "", regex=True)
    out["Sumilla"] = out["Sumilla"].str.strip().replace("", pd.NA)

    # 3. Columnas derivadas
    out["uuid"] = out["Link Resolucion"].str.extract(_UUID, expand=False)
    fecha = pd.to_datetime(out["Fecha de Resolucion"], format="%d/%m/%Y", errors="coerce")
    out["Fecha"] = fecha
    out["Anio"] = fecha.dt.year.astype("Int64")
    anio_exp = pd.to_numeric(
        out["Nro Expediente"].str.extract(r"-(\d{4})$", expand=False), errors="coerce"
    ).astype("Int64")
    out["Anio Expediente"] = anio_exp

    # 4. Banderas de calidad
    out["sumilla_tenia_prefijo"] = tenia_prefijo
    out["sumilla_util"] = out["Sumilla"].str.len().ge(LARGO_MINIMO_SUMILLA).fillna(False)
    out["fecha_incoherente"] = (out["Anio"].notna() & anio_exp.notna() & (out["Anio"] < anio_exp))

    # uuid repetido (colision del portal): dos filas apuntando al mismo PDF, asi
    # que una de las dos tendra metadatos que no corresponden al documento.
    out["uuid_repetido"] = out["uuid"].duplicated(keep=False) & out["uuid"].notna()

    # 5. Ruta de destino del PDF
    out["Archivo PDF"] = [
        ruta_pdf(r, e, esp, u, a)
        for r, e, esp, u, a in zip(
            out["Casacion/Apelacion"], out["Nro Expediente"], out["Especialidad"],
            out["uuid"], out["Anio"]
        )
    ]
    # El markdown espeja la ruta del PDF: unir texto y metadatos es cambiar la
    # extension, sin tabla intermedia ni riesgo de desparejar.
    out["Archivo MD"] = out["Archivo PDF"].str.replace(r"\.pdf$", ".md", regex=True)

    # Casi-duplicados: mismo contenido, distinto uuid. Se numera cada grupo para
    # poder inspeccionarlos juntos; NA en las filas que no repiten nada.
    llave = out[CONTENIDO].fillna("\x00").agg("\x1f".join, axis=1)
    tam = llave.map(llave.value_counts())
    grupo = llave.where(tam > 1).astype("category").cat.codes
    out["grupo_duplicado"] = pd.Series(grupo, index=out.index).where(tam > 1).astype("Int64")
    out["es_duplicado_contenido"] = tam > 1
    # El primero de cada grupo es el representante; el resto son las copias.
    out["copia_descartable"] = out["es_duplicado_contenido"] & out.duplicated(
        subset=CONTENIDO, keep="first"
    )

    return out[ORIGINALES + [c for c in out.columns if c not in ORIGINALES]]


def verificar(antes, despues):
    """Comprueba que no se ha perdido, añadido ni movido nada. Devuelve errores."""
    fallos = []

    if len(antes) != len(despues):
        fallos.append(f"nº de filas: {len(antes):,} -> {len(despues):,}")

    if not antes.index.equals(despues.index):
        fallos.append("el indice ha cambiado")

    faltan = [c for c in ORIGINALES if c not in despues.columns]
    if faltan:
        fallos.append(f"faltan columnas originales: {faltan}")

    if list(despues.columns[: len(ORIGINALES)]) != ORIGINALES:
        fallos.append("las columnas originales no estan en su orden")

    # El orden de las filas: la secuencia de uuid debe ser identica. Se compara
    # valor a valor y no con .equals(), que exige ademas el mismo dtype y nombre.
    u_antes = antes["Link Resolucion"].str.extract(_UUID, expand=False).astype("string")
    distintos = int((u_antes.fillna("\x00") != despues["uuid"].fillna("\x00")).sum())
    if distintos:
        fallos.append(f"la secuencia de uuid no coincide ({distintos:,} posiciones)")

    # Ningun campo debe haberse vaciado. Unica excepcion admitida: sumillas que
    # solo contenian la etiqueta o su puntuacion, sin texto detras.
    for col in ORIGINALES:
        nuevos = antes[col].notna() & despues[col].isna()
        if not nuevos.any():
            continue
        if col == "Sumilla":
            resto = (
                antes.loc[nuevos, col]
                .astype("string")
                .str.replace(_LIMPIEZA_INICIAL, "", regex=True)
                .str.strip()
            )
            con_texto = int((resto.fillna("") != "").sum())
            if con_texto:
                fallos.append(f"'Sumilla': {con_texto:,} valores con texto pasaron a nulo")
            continue
        fallos.append(f"'{col}': {int(nuevos.sum()):,} valores pasaron a nulo")

    # Postcondiciones de la limpieza de texto: si algo quedo a medias, se ve aqui.
    sumillas = despues["Sumilla"].dropna()
    resto = int(sumillas.str.match(_LIMPIEZA_INICIAL).sum())
    if resto:
        fallos.append(f"{resto:,} sumillas conservan la etiqueta o puntuacion inicial")
    for col in TEXTO:
        sucias = int(despues[col].dropna().str.contains(r"\s\s|[\n\r\t]", regex=True).sum())
        if sucias:
            fallos.append(f"'{col}': {sucias:,} celdas con espaciado sin limpiar")

    # Lo que sostiene todo el paso de descarga: una fila, un uuid, una ruta.
    uuid = despues["uuid"]
    if uuid.isna().any():
        fallos.append(f"{int(uuid.isna().sum()):,} filas sin uuid")
    malos = int((~uuid.fillna("").str.fullmatch(r"[0-9a-f-]{36}")).sum())
    if malos:
        fallos.append(f"{malos:,} uuid con formato invalido")
    if not uuid.equals(despues["Link Resolucion"].str.extract(_UUID, expand=False).astype("string")):
        fallos.append("el uuid derivado no coincide con el del link")

    rutas = despues["Archivo PDF"]
    repetidas = int(len(rutas) - rutas.nunique())
    if repetidas:
        fallos.append(f"{repetidas:,} rutas de PDF colisionan (se sobrescribirian)")
    if not rutas.str.fullmatch(r"[a-z0-9-]+/(?:[0-9]{4}|sin-anio)/[a-z0-9-]+\.pdf").all():
        fallos.append("hay rutas de PDF con caracteres inesperados")

    # Las columnas que no tocamos deben ser identicas.
    for col in ("Nro Expediente", "Fecha de Resolucion", "Link Resolucion", "Pagina"):
        a = antes[col].astype("string").str.strip().replace("", pd.NA)
        b = despues[col].astype("string")
        if not a.fillna("\x00").equals(b.fillna("\x00")):
            fallos.append(f"'{col}' fue modificada y no deberia")

    return fallos


def informe(antes, despues):
    def linea(t, v):
        print(f"  {t:.<52} {v}")

    print("\nCATEGORIAS UNIFICADAS")
    for col in ("Casacion/Apelacion", "Tipo de Resolucion", "Sala Suprema", "Especialidad"):
        a, b = antes[col].nunique(), despues[col].nunique()
        marca = "" if a == b else f"   <-- {a - b} variante(s) colapsada(s)"
        linea(col, f"{a} -> {b}{marca}")
        if a != b:
            cambiadas = antes.loc[antes[col] != despues[col], col].dropna().unique()
            for v in cambiadas[:10]:
                destino = despues.loc[antes[col] == v, col].iloc[0]
                print(f"      {v!r} -> {destino!r}")

    print("\nTEXTO LIMPIADO")
    for col in TEXTO:
        n = int((antes[col].fillna("") != despues[col].fillna("")).sum())
        linea(col, f"{n:,} celdas retocadas")

    print("\nBANDERAS")
    linea("sumillas con prefijo 'Sumilla:' quitado", f"{int(despues['sumilla_tenia_prefijo'].sum()):,}")
    vaciadas = int((antes["Sumilla"].notna() & despues["Sumilla"].isna()).sum())
    linea("sumillas que solo eran la etiqueta (-> nulo)", f"{vaciadas:,}")
    linea("sumillas utiles (>=20 caracteres)", f"{int(despues['sumilla_util'].sum()):,}")
    linea("fechas anteriores al año del expediente", f"{int(despues['fecha_incoherente'].sum()):,}")
    linea("filas con uuid repetido", f"{int(despues['uuid_repetido'].sum()):,}")
    linea("filas con contenido duplicado", f"{int(despues['es_duplicado_contenido'].sum()):,}")
    linea("  ...en nº de grupos", f"{int(despues['grupo_duplicado'].nunique()):,}")
    linea("  ...copias descartables", f"{int(despues['copia_descartable'].sum()):,}")
    linea("filas unicas si se quitan las copias", f"{len(despues) - int(despues['copia_descartable'].sum()):,}")

    print("\nDEDUPLICACION DE SUMILLAS (efecto de quitar el prefijo)")
    a = antes["Sumilla"].dropna().nunique()
    b = despues["Sumilla"].dropna().nunique()
    linea("sumillas distintas", f"{a:,} -> {b:,}   ({a - b:,} eran la misma)")

    print("\nDESCARGA DE PDFs")
    sin_uuid = (
        despues["Casacion/Apelacion"].fillna("resolucion").map(_slug)
        + "-"
        + despues["Nro Expediente"].fillna("")
    )
    perdidos = len(despues) - sin_uuid.nunique()
    linea("nombres 'recurso-expediente' distintos", f"{sin_uuid.nunique():,}")
    linea("  ...PDFs que se perderian con ese esquema", f"{perdidos:,}")
    linea("rutas 'Archivo PDF' distintas", f"{despues['Archivo PDF'].nunique():,}")
    linea("uuid distintos a descargar", f"{despues['uuid'].nunique():,}")
    carpetas = despues["Archivo PDF"].str.rsplit("/", n=1).str[0]
    linea("carpetas de destino", f"{carpetas.nunique()}")
    linea("  ...ficheros en la mayor", f"{carpetas.value_counts().iloc[0]:,}")
    linea("longitud maxima de ruta", f"{int(despues['Archivo PDF'].str.len().max())} caracteres")
    comp = int(despues["uuid_repetido"].sum())
    if comp:
        linea("filas que comparten PDF (revisar a mano)", f"{comp}")


def escribir(df, destino):
    """Excel con cabecera fija y autofiltro, en modo write_only por tamaño."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    libro = Workbook(write_only=True)
    hoja = libro.create_sheet("Resoluciones")
    hoja.freeze_panes = "A2"

    anchos = {"Sumilla": 70, "Link Resolucion": 62, "Palabras Clave": 34,
              "Pretension/Delito": 34, "Sala Suprema": 34}
    for i, col in enumerate(df.columns, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = anchos.get(col, 18)

    negrita = Font(bold=True)
    from openpyxl.cell import WriteOnlyCell
    cabecera = []
    for col in df.columns:
        celda = WriteOnlyCell(hoja, value=col)
        celda.font = negrita
        celda.alignment = Alignment(vertical="center")
        cabecera.append(celda)
    hoja.append(cabecera)

    # Excel no distingue booleano de 0/1 si le llega un numpy.bool_, y al releer
    # el fichero saldrian enteros: se pasan a bool de Python para que queden
    # como VERDADERO/FALSO y sigan sirviendo de mascara al recargarlos.
    salida = df.copy()
    for col in salida.columns:
        if pd.api.types.is_bool_dtype(salida[col]):
            salida[col] = salida[col].fillna(False).map(bool).astype(object)
    # Fecha como fecha real, no como texto, para poder filtrar por rango.
    salida["Fecha"] = salida["Fecha"].dt.date

    for fila in salida.itertuples(index=False, name=None):
        hoja.append([None if pd.isna(v) else v for v in fila])

    destino.parent.mkdir(parents=True, exist_ok=True)
    libro.save(destino)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("entrada", nargs="?", type=Path, default=ENTRADA)
    ap.add_argument("salida", nargs="?", type=Path, default=None)
    ap.add_argument("--solo-informe", action="store_true", help="no escribe ningun fichero")
    args = ap.parse_args()

    if not args.entrada.exists():
        print(f"no existe: {args.entrada}", file=sys.stderr)
        return 1
    destino = args.salida or args.entrada.with_name(args.entrada.stem + "-limpio.xlsx")

    print(f"leyendo {args.entrada}")
    antes = pd.read_excel(args.entrada, engine="openpyxl")
    print(f"  {len(antes):,} filas x {len(antes.columns)} columnas")

    despues = limpiar(antes)
    informe(antes, despues)

    print("\nVERIFICACION")
    fallos = verificar(antes, despues)
    if fallos:
        for f in fallos:
            print(f"  FALLO: {f}")
        print("\n=> NO se escribe nada. El original sigue intacto.")
        return 1
    print("  filas, orden, indice y columnas originales: intactos")
    print("  ningun valor se perdio, ninguna fila se movio")

    if args.solo_informe:
        print("\n(--solo-informe: no se escribe nada)")
        return 0

    lock = destino.parent / f".~lock.{destino.name}#"
    if lock.exists():
        print(f"\n{destino.name} esta abierto en LibreOffice, cierralo primero.", file=sys.stderr)
        return 1

    manifiesto = destino.with_name("manifiesto-pdfs.csv")
    print(f"\nescribiendo {manifiesto}")
    despues.rename(columns={"Link Resolucion": "url", "Archivo PDF": "ruta",
                            "Archivo MD": "ruta_md"})[
        ["uuid", "url", "ruta", "ruta_md", "Nro Expediente", "Fecha", "Anio",
         "Especialidad", "uuid_repetido"]
    ].to_csv(manifiesto, index=False)
    print(f"  {len(despues):,} filas, {despues['uuid'].nunique():,} PDFs distintos")

    print(f"\nescribiendo {destino}")
    escribir(despues, destino)
    print(f"  {len(despues):,} filas x {len(despues.columns)} columnas")
    print(f"  original intacto: {args.entrada}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
