"""Mete en el corpus un Excel cosechado por fuera, verificandolo primero.

larijo cosecho su parte con codigo propio y solo tiene los Excel: ni tramos ni
uuid. No pasa nada, porque el Excel trae lo necesario para reconstruirlos. El
uuid va dentro del propio link (ServletDescarga?uuid=...), que es como lo
armo juris_parse en su momento, y la columna Pagina permite buscar huecos igual
que hace auditar.py con los tramos.

Como esos datos no salieron de este codigo, no se dan por buenos: se revisa
que las columnas esten, que los links tengan forma de link, que no falten
paginas y que ninguna traiga mas de 10 resoluciones. Solo si sale limpio se
escriben los tramos.

Uso:
    python3 importar.py salida-larijo.xlsx              # solo revisa
    python3 importar.py *.xlsx --escribir               # revisa y mete
"""

import argparse
import collections
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from juris_client import ESPECIALIDADES, REPARTO
from juris_excel import COLUMNAS, SALIDA, slug

TRAMO = 25
POR_PAGINA = 10
_UUID_EN_LINK = re.compile(r"ServletDescarga\?uuid=([0-9a-fA-F-]{36})")


def leer_excel(ruta):
    """Filas del Excel como diccionarios, con el uuid sacado del link."""
    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active
    iterador = hoja.iter_rows(values_only=True)

    try:
        cabecera = [str(c).strip() if c is not None else "" for c in next(iterador)]
    except StopIteration:
        raise ValueError("el fichero esta vacio")

    faltan = [c for c in COLUMNAS if c not in cabecera]
    if faltan:
        raise ValueError(f"le faltan columnas: {', '.join(faltan)}")

    indice = {c: cabecera.index(c) for c in COLUMNAS}
    filas, sin_link, link_raro = [], 0, 0

    for cruda in iterador:
        if cruda is None or all(c is None for c in cruda):
            continue
        fila = {c: (cruda[indice[c]] if cruda[indice[c]] is not None else "")
                for c in COLUMNAS}

        link = str(fila["Link Resolucion"]).strip()
        if not link:
            sin_link += 1
            continue
        m = _UUID_EN_LINK.search(link)
        if not m:
            link_raro += 1
            continue

        fila["uuid"] = m.group(1)
        fila["Link Resolucion"] = link
        # Igual que juris_parse marca si una fila salio del payload JSON o del
        # HTML, aqui queda constancia de que no la cosecho este proyecto. Para
        # el dataset importa saber que cuarto del corpus vino de fuera.
        fila["Origen"] = "excel"
        try:
            fila["Pagina"] = int(fila["Pagina"])
        except (TypeError, ValueError):
            fila["Pagina"] = 0
        filas.append(fila)

    return filas, sin_link, link_raro


def revisar(filas):
    """Agrupa por especialidad y saca los problemas de cada una."""
    por_esp = collections.defaultdict(list)
    for fila in filas:
        por_esp[str(fila["Especialidad"]).strip()].append(fila)

    informe = {}
    for esp, suyas in por_esp.items():
        paginas = collections.Counter(f["Pagina"] for f in suyas)
        uuids = {f["uuid"] for f in suyas}
        ultima = max(paginas) if paginas else 0
        informe[esp] = {
            "filas": suyas,
            "unicas": len(uuids),
            "duplicadas": len(suyas) - len(uuids),
            "paginas": len(paginas),
            "ultima": ultima,
            "faltan": sorted(set(range(1, ultima + 1)) - set(paginas)),
            "sobrecargadas": sorted(p for p, n in paginas.items() if n > POR_PAGINA),
            "sin_pagina": paginas.get(0, 0),
            "conocida": esp in ESPECIALIDADES,
        }
    return informe


def ya_cosechada(esp):
    """Si esa especialidad ya tiene tramos propios en salida/."""
    return any((SALIDA / slug(esp) / "tramos").glob("p*.json"))


def escribir_tramos(esp, datos):
    """Deja las filas como tramos, con el formato que espera el resto del proyecto."""
    carpeta = SALIDA / slug(esp) / "tramos"
    carpeta.mkdir(parents=True, exist_ok=True)

    por_pagina = collections.defaultdict(list)
    for fila in datos["filas"]:
        por_pagina[fila["Pagina"]].append(fila)

    escritos = 0
    for inicio in range(1, datos["ultima"] + 1, TRAMO):
        fin = min(inicio + TRAMO - 1, datos["ultima"])
        filas = [f for p in range(inicio, fin + 1) for f in por_pagina.get(p, [])]
        if not filas:
            continue
        faltan = [p for p in range(inicio, fin + 1) if p not in por_pagina]
        destino = carpeta / f"p{inicio:05d}-{fin:05d}.json"
        destino.write_text(
            json.dumps({"completo": not faltan, "fallidas": faltan, "filas": filas},
                       ensure_ascii=False),
            encoding="utf-8",
        )
        escritos += 1
    return escritos


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("ficheros", nargs="+", help="los .xlsx a importar")
    ap.add_argument("--escribir", action="store_true",
                    help="escribe los tramos en salida/ (sin esto solo revisa)")
    ap.add_argument("--persona", choices=sorted(REPARTO),
                    help="importa solo las especialidades de esa persona")
    ap.add_argument("--forzar", action="store_true",
                    help="importa aunque la especialidad ya tenga tramos cosechados")
    args = ap.parse_args()

    todas, problemas = [], []
    for ruta in args.ficheros:
        try:
            filas, sin_link, link_raro = leer_excel(Path(ruta))
        except Exception as e:
            print(f"{ruta}: NO SE PUDO LEER -> {e}")
            problemas.append(ruta)
            continue
        print(f"{ruta}: {len(filas)} filas con link")
        if sin_link:
            print(f"  !! {sin_link} filas sin link, descartadas")
            problemas.append(ruta)
        if link_raro:
            print(f"  !! {link_raro} links sin uuid dentro, descartados")
            problemas.append(ruta)
        todas.extend(filas)

    if not todas:
        print("\nNo hay nada que importar.")
        return 1

    informe = revisar(todas)

    # Un Excel puede traer especialidades que no son las que se quieren meter:
    # los consolidados llevan las cuatro de su dueno, y un glob suelto acaba
    # mezclando los de todo el equipo.
    if args.persona:
        suyas = set(REPARTO[args.persona])
        fuera = sorted(e for e in informe if e not in suyas)
        if fuera:
            print(f"\nNo son de {args.persona}, las ignoro: {', '.join(fuera)}")
        informe = {e: d for e, d in informe.items() if e in suyas}
        if not informe:
            print(f"\nNinguna de las especialidades de {args.persona} en esos ficheros.")
            return 1

    print(f"\n{'=' * 60}")

    for esp in sorted(informe):
        d = informe[esp]
        malo = (d["faltan"] or d["duplicadas"] or d["sobrecargadas"]
                or d["sin_pagina"] or not d["conocida"])
        print(f"\n{esp}  [{'REVISAR' if malo else 'OK'}]")
        print(f"  {d['unicas']} resoluciones unicas en {d['paginas']} paginas "
              f"(hasta la {d['ultima']})")
        if not d["conocida"]:
            print(f"  !! no es una especialidad del proyecto")
        if d["duplicadas"]:
            print(f"  !! {d['duplicadas']} filas con uuid repetido")
        if d["sin_pagina"]:
            print(f"  !! {d['sin_pagina']} filas sin numero de pagina valido")
        if d["sobrecargadas"]:
            print(f"  !! {len(d['sobrecargadas'])} paginas con mas de {POR_PAGINA} filas: "
                  f"{d['sobrecargadas'][:10]}")
        if d["faltan"]:
            print(f"  !! faltan {len(d['faltan'])} paginas: "
                  f"{d['faltan'][:40]}{' ...' if len(d['faltan']) > 40 else ''}")
        if malo:
            problemas.append(esp)

    total = sum(d["unicas"] for d in informe.values())
    print(f"\n{'=' * 60}")
    print(f"TOTAL: {total} resoluciones unicas en {len(informe)} especialidades")

    if problemas:
        print("\nHay cosas que revisar. No se escribe nada hasta resolverlas.")
        print("Si aun asi quieres meterlo, vuelve a lanzarlo con --escribir:")
        print("los tramos con paginas faltantes quedaran marcados como incompletos")
        print("y la cosecha los rehara.")

    if not args.escribir:
        print("\nSolo he revisado. Para meterlo en salida/:  --escribir")
        return 1 if problemas else 0

    print()
    for esp, d in sorted(informe.items()):
        if not d["conocida"]:
            print(f"  {esp}: saltada, no es del proyecto")
            continue
        # Lo cosechado aqui vale mas que un Excel: viene con los checkpoints y
        # ha pasado por auditar.py. Un consolidado viejo de la propia maquina
        # trae los numeros de mitad de la corrida, asi que sobrescribir seria
        # cambiar datos buenos por datos peores sin enterarse.
        if ya_cosechada(esp) and not args.forzar:
            print(f"  {esp}: SALTADA, ya tiene tramos cosechados (--forzar para pisarlos)")
            continue
        n = escribir_tramos(esp, d)
        print(f"  {esp}: {n} tramos escritos en salida/{slug(esp)}/tramos/")

    print("\nHecho. Comprueba con:  python3 auditar.py")
    return 1 if problemas else 0


if __name__ == "__main__":
    sys.exit(main())
