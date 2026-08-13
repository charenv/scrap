"""Escritura del Excel y rutas de salida, compartido por cosecha.py y exportar.py."""

import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SALIDA = Path(__file__).parent / "salida"

# Las columnas del README, en orden. Todo Excel del proyecto usa estas mismas.
COLUMNAS = [
    "Casacion/Apelacion",
    "Nro Expediente",
    "Pretension/Delito",
    "Tipo de Resolucion",
    "Fecha de Resolucion",
    "Sala Suprema",
    "Norma de Derecho Interno (Articulo)",
    "Sumilla",
    "Palabras Clave",
    "Link Resolucion",
    "Especialidad",
    "Pagina",
]
_ANCHOS = [18, 16, 34, 20, 16, 34, 30, 70, 34, 62, 22, 8]


def slug(texto):
    """'Contencioso Adm. Laboral' -> 'contencioso-adm-laboral'."""
    plano = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", plano.lower()).strip("-")


def escribir_excel(filas, destino):
    """Genera el Excel con las columnas del README, listo para filtrar."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Resoluciones"

    hoja.append(COLUMNAS)
    for celda in hoja[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(vertical="center")

    for fila in filas:
        hoja.append([fila.get(columna, "") for columna in COLUMNAS])

    for i, ancho in enumerate(_ANCHOS, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    destino.parent.mkdir(parents=True, exist_ok=True)
    libro.save(destino)
